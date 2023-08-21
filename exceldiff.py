import openpyxl

def cruzarDatos(pedido_alcance, pedido_analista, doc_salida):
        wb_analista = openpyxl.load_workbook(pedido_analista)
        wb_alcance=openpyxl.load_workbook(pedido_alcance)
        wb_output=openpyxl.Workbook()


        ws_analista=wb_analista.active
        ws_alcance=wb_alcance.active
        ws_output=wb_output.active

        guia_analista=0
        guia_alcance=0

        id_item_analista=-1
        cantidad_analista=0
        id_item_alcance=0
        cantidad_alcance=0
        descripcion_alcance=''
        descripcion_analista=''
        fila_libro=0
        flujo_control_encontrar_id_item=0
        for row_alcance in ws_alcance.iter_rows():
            for cell_alcance in row_alcance:
                guia_alcance=guia_alcance+1

                if(cell_alcance.value is None):
                    continue
                if((guia_alcance % 3) ==1):
                    id_item_alcance=int(str(cell_alcance.value))

                if((guia_alcance % 3) ==2):

                    descripcion_alcance=str(cell_alcance.value)
                if((guia_alcance % 3) ==0):
                    cantidad_alcance=int(str(cell_alcance.value))
                flujo_control_encontrar_id_item=0
                for row_analista in ws_analista.iter_rows():

                    for cell_analista in row_analista:
                            guia_analista=guia_analista+1
                            if(cell_analista.value is None):
                                continue
                            if((guia_analista % 3 )==1):
                                id_item_analista=int(str(cell_analista.value))
                            if((guia_analista % 3 )== 2):
                                    descripcion_analista=str(cell_analista.value)
                            if((guia_analista % 3 )== 0):
                                cantidad_analista=int(str(cell_analista.value))
                    if(id_item_alcance==id_item_analista):

                        #if(cantidad_alcance>cantidad_analista):
                   #                 fila_libro=fila_libro+1
                        cantidad_alcance=cantidad_alcance-cantidad_analista
                #                    ws_output['A' + str(fila_libro)]=str(id_item_alcance)
                 #                   ws_output['B' + str(fila_libro)]=str(descripcion_analista)
                  #                  ws_output['C' + str(fila_libro)]=str(cantidad_alcance)
                        flujo_control_encontrar_id_item=1

                        print(str(id_item_alcance) + "   "+   str(descripcion_analista)  + "   "  + str(cantidad_alcance))
                    if(flujo_control_encontrar_id_item==1):
                        id_item_analista=-1
                        break



                #if(flujo_control_encontrar_id_item==0):
            if(cantidad_alcance>0):
                fila_libro=fila_libro+1
                ws_output['A' +  str(fila_libro)]=str(id_item_alcance)
                ws_output['B' +  str(fila_libro)]=str(descripcion_alcance)
                ws_output['C' +  str(fila_libro)]=str(cantidad_alcance)


        wb_output.save(str(doc_salida) + '.xlsx')