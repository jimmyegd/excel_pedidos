import openpyxl
import sqlite3
import math





def alcance(curvaABC, dias_analisis_arg, dias_alcance_arg, pedidoAlcance):

    # Load the workbook
    wb = openpyxl.load_workbook(curvaABC)
    wb_output=openpyxl.Workbook()
    conn = sqlite3.connect('sqlite/pedidospy.db')



    # Select the active worksheet
    ws = wb.active
    ws_output=wb_output.active
    guia=0

    id_item=0
    num_fraccion=0
    Cant_U=0
    Cant_F=0
    U_Ven=0
    F_Ven=0
    total_f_stock=0
    total_f_ven=0
    dias_analisis=int(dias_analisis_arg)
    dias_alcance=int(dias_alcance_arg)
    venta_alcance=0.0
    alcance_unidades=0.0
    descripcion=""
    # Access the values of cells in the worksheet
    control=0
    rows_1=0

    for row in ws.iter_rows():
        
        for cell in row:
            guia=guia+1
            if(cell.value is None):
                    continue
            if((guia % 6)==1):
                id_item=str(cell.value)
                #print(str(cell.value)  + "  Id_item  "  )
                cursor = conn.execute("SELECT num_fraccion from productos where id_producto=" + str(cell.value))
                res=cursor.fetchone()
                control=0
                if( res is not None):
                    num_fraccion=int(res[0])
                 #   print(str(res[0]) + " num_fraccion")
                else:
                    control=1
                    print("No existe num_fraccion para el producto de id=" + str(cell.value)   )
            
                
            if((guia % 6)==2):
                descripcion=str(cell.value)    
                #print( descripcion + "  descripcion  "  )
            if((guia % 6)==3):
                Cant_U=int(str(cell.value))
                #print(str(cell.value)  + "  Cant_U  "  )
            if((guia % 6)==4):
                Cant_F=int(str(cell.value))
                #print(str(cell.value)  + "  Cant_F  "  )
            if((guia % 6)==5):
                U_Ven=int(str(cell.value))
                #print(str(cell.value)  + "  U_Ven  "  )
            if((guia % 6)==0):
                #print(str(cell.value)  + "  F_Ven  "  + str(guia))
                F_Ven=int(str(cell.value))
                
                
                total_f_stock=(Cant_U*num_fraccion)+Cant_F
                total_f_ven=(U_Ven*num_fraccion)+F_Ven
                venta_alcance=float(total_f_ven)*float(dias_alcance)/float(dias_analisis)
                #print("Total stock :  " +str(total_f_stock))
                #print("Total venta :  " +str(total_f_ven))
                #print("Total venta_alcance :  " +str(venta_alcance))
                alcance_unidades=(float(venta_alcance)-float(total_f_stock))/float(num_fraccion)
                #print("Total alcance_unidades :  " +str(alcance_unidades))
                if(alcance_unidades>0  and control ==0):
                    rows_1=rows_1+1
                    pedir=math.ceil(alcance_unidades)
                    ws_output['A'+str(rows_1)]=str(id_item)
                    ws_output['B'+str(rows_1)]=str(descripcion)
                    ws_output['C'+ str(rows_1)]=str(pedir)
                    print(str(id_item) + "  : "   +   str(descripcion) + "  : "  +  str(pedir))
                
                
                
                
            
     


    wb_output.save(str(pedidoAlcance)+'.xlsx') 
    print(str(pedidoAlcance))
    conn.close()            